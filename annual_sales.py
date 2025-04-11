import pandas as pd
from collections import OrderedDict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill

from annual_sales_calculator import Calculator
from annual_sales_DataLoader import DataLoader





# %% Class [ sheet_name = 114年 ]
# =============================================================================
# StoreReport: 負責讀取各家店原始資料、計算各項指標
# =============================================================================
class StoreReport:
    def __init__(self, Config):
        self.Config = Config
        self.year = Config['year']
        self.month = Config['month']
        self.company_name = Config['company_name']
        self.each_area_path = Config['each_area_path']
        self.path_one = Config['path_one']
        self.path_two = Config['path_two']
        self.last_year_path = Config['last_year_path']
        self.this_year_path = Config['this_year_path']
        self.path_four = Config['path_four']
        self.path_five = Config['path_five']
        self.store_name_list, self.total_store_number = self._get_store_name()
        self.all_data_dict = {store_name: {m+1: {} for m in range(self.month)} for store_name in self.store_name_list}

        #後續會用到的參數
        self.data_dict = None
        self.total_sales = None
        self.final_data_dict = None
        self.final_df = None
          
    
    def build_StoreReport_all(self):
        '''建好所有Data，從Excel -> dict -> DataFrame -> 加入計算'''
        #取得 -> all_data_dict
        self.get_all_excel_data_to_dict()
        #取得 -> data_dict
        self.convert_dict_to_df()
        #取得 -> final_data_dict, final_df
        self.cal_other_index_value()   
        
        
    #每一家店名稱
    def _get_store_name(self): 
        '''生成所有店的店名'''
        df = pd.read_excel(self.each_area_path, sheet_name = '區域分店')
        self.store_name_list = list(df.iloc[:, 1])
        self.total_store_number = len(self.store_name_list) 

        return self.store_name_list, self.total_store_number
    
    def _build_main_df(self):    
        '''
        建立 營業店年度營業額及各項比率計算(114年).xlsx 所需之df
    
        '''
        self.index_list = ['銷貨收入','折扣金額','折扣佔比','營業目標','營業額','營業目標達成率','營業額佔比','去年營業額',
                      '成長率','費用','費用率','毛利','毛利率','實際毛利率','淨利目標','淨利達成率','單位淨利','淨利率',
                      '應發薪資','薪資比','奬金','奬金比','薪資生產力','工時合計','工時生產力','來客數','客單價']
        
        self.columns_list = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十', '十一', '十二', '合  計']    

        # 建立空的 DataFrame
        df = pd.DataFrame(index=self.index_list, columns=self.columns_list)    
        
        return df
    

    
    #### 1. 讀取的資料存到 -> all_data_dict
    
    def get_all_excel_data_to_dict(self):
        data_loader = DataLoader(self.Config, self.store_name_list, self.all_data_dict, center_kitchen_df = None)

        self.all_data_dict = data_loader.load_all()
        self.rent_df = data_loader.rent_df
        
        return self.all_data_dict
        
    #### 2. 讀取的資料轉成 DataFrame -> data_dict
    def convert_dict_to_df(self):
        '''
        將所有從excel讀取的資料輸入到 dataframe
        依據各家店名稱 儲存成一個dict
        且計算單一家店最右邊合計
        data_dict = { 店名 : { dataframe } }
    
        '''
        self.data_dict = {store_name:{} for store_name in self.store_name_list}
 
        for name in self.store_name_list:
            df = self._build_main_df()
            for i in range(self.month):
                if i+1 in self.all_data_dict[name].keys():
                    for index_name, value in self.all_data_dict[name][i+1].items():
                        if index_name in self.index_list:
                            df.loc[index_name, df.columns[i]] = value
                                         
            #單一家店最右邊合計
            row_sums = df.sum(axis=1)
            df.iloc[:, -1] = row_sums
            self.data_dict[name] = df


    # 計算各個項目輸入dataframe 
    @staticmethod       
    def _cal_total_sale_at_numerator(df):
        #營業目標達成率 = 營業額 / 營業目標
        Calculator._cal_division(df, '營業額', '營業目標', '營業目標達成率')        
        #薪資生產力 = 營業額 / 應發薪資
        Calculator._cal_division(df, '營業額', '應發薪資', '薪資生產力')        
        #工時生產力 = 營業額 / 工時合計
        Calculator._cal_division(df, '營業額', '工時合計', '工時生產力')        
        #客單價 = 營業額 / 來客數
        Calculator._cal_division(df, '營業額', '來客數', '客單價')

    @staticmethod
    def _cal_total_sales_at_denominator(df):
        #費用率 = 費用 / 營業額
        Calculator._cal_division(df, '費用', '營業額', '費用率')        
        #毛利率 = 毛利 / 營業額
        Calculator._cal_division(df, '毛利', '營業額', '毛利率')        
        #淨利率 = 單位淨利 / 營業額
        Calculator._cal_division(df, '單位淨利', '營業額', '淨利率')        
        #薪資比 = 應發薪資 / 營業額
        Calculator._cal_division(df, '應發薪資', '營業額', '薪資比')       
        #獎金比 = 獎金 / 營業額
        Calculator._cal_division(df, '奬金', '營業額', '奬金比')
        
    
    ####3. 計算其他index數值
    def cal_other_index_value(self):
        '''
        計算 dataframe 中需要計算的 index
    
        '''       
        #總營業額
        self.total_sales = sum(df.loc['營業額'].fillna(0) for df in self.data_dict.values())
        
        #存放最終結果的dict    
        self.final_data_dict = OrderedDict()
        # name = '新田店'
        for name in self.store_name_list:
            df = self.data_dict[name]
            #銷貨收入 = 折扣金額 + 營業額
            Calculator._cal_sales_revenue(df)            
            #折扣佔比 = 折扣金額 / 銷貨收入
            Calculator._cal_division(df, '折扣金額', '銷貨收入', '折扣佔比')            
            #營業額佔比 = 營業額 / 總營業額
            Calculator._cal_sales_proportion(self.total_sales, df, '營業額佔比')            
            #成長率 = (營業額 - 去年營業額) / 去年營業額
            Calculator._cal_growth_rate(df)    
            #淨利達成率 = 單位淨利 / 淨利目標
            Calculator._cal_division(df, '單位淨利', '淨利目標', '淨利達成率')
            #營業額當分子
            StoreReport._cal_total_sale_at_numerator(df)
            #營業額當分母
            StoreReport._cal_total_sales_at_denominator(df)
                
            self.final_data_dict[name] = df
        
        #最終的df
        self.final_df = pd.concat(self.final_data_dict)
        
        
        return self.final_data_dict, self.final_df
        



# %% Class 所有店家總和 + 各類別店家總和
class StoreSummary(StoreReport):
    def __init__(self, Config):
        super().__init__(Config) 
        self.store_type_list = None
        self.store_type_dict = None
        self.concat_all_type_store_df = pd.DataFrame()
        self.total_df = None
    
    def build_StoreSummary_all(self):
        self.build_StoreReport_all()
        self._get_store_type()
        self.cal_each_type_store()
        final_result_df = self.concat_all_type_store_sum()
        self._cal_same_period_profit()
        return final_result_df
    
    def _cal_total_sum_df(self, data_dict):

        ### 將所有店家的 dataframe相加並未計算需要用公式的欄位
        # 取得字典中的第一個DataFrame作為起點
        df = data_dict[list(data_dict.keys())[0]].copy()
        
        # 從第二個DataFrame開始相加()
        for key in list(data_dict.keys())[1:]:
            df = df.add(data_dict[key], fill_value=0)
                  
        ### 計算需要用公式的欄位                     
        #折扣佔比 = 折扣金額 / 銷貨收入
        Calculator._cal_division(df, '折扣金額', '銷貨收入', '折扣佔比')        
        #營業額佔比 = 營業額 / 總營業額
        Calculator._cal_sales_proportion(self.total_sales, df, '營業額佔比')
        #成長率 = (營業額 - 去年營業額) / 去年營業額
        Calculator._cal_growth_rate(df)    
        #淨利達成率 = 單位淨利 / 淨利目標
        Calculator._cal_division(df, '單位淨利', '淨利目標', '淨利達成率')        
        #營業額當分子
        StoreReport._cal_total_sale_at_numerator(df)
        #營業額當分母
        StoreReport._cal_total_sales_at_denominator(df)
        
        return df
        
        
    
    #### 各家店類別 -> 店中店 or 美食街    
    def _get_store_type(self):
        '''
        store_type_dict : { 店家類型 : 該類型之店家 }
    
        '''
        try:
            df = pd.read_excel(self.each_area_path, sheet_name='分店類別')
        except FileNotFoundError:
            raise FileNotFoundError(f"找不到指定的檔案：{self.each_area_path}，請確認路徑正確！")
        except ValueError as e:
            # 通常是找不到指定的sheet name
            raise ValueError(f"讀取Excel失敗：{e}，請確認 '分店類別' 工作表存在！")
        except Exception as e:
            raise Exception(f"讀取Excel時發生未知錯誤：{e}")
        
        #店中店、美食街
        self.store_type_list = list(set(df.iloc[:, 0]))
        
        #各類別店家
        self.store_type_dict = OrderedDict()
        for store_type in self.store_type_list:
            individual_type_df = list(df.loc[df.iloc[:, 0] == store_type].iloc[:, 1])
            self.store_type_dict[store_type] = individual_type_df
        
        return self.store_type_dict
    
    #店中店及美食街的dict
    def cal_each_type_store(self):
        '''
        從先前計算完畢的 final_data_dict 中
        取得各分店類別的 dataframe
        all_store_type_dict = { 分店類別 ->  各分店 -> dataframe } 
    
        '''
        self.total_df = self._cal_total_sum_df(self.final_data_dict)
        
        #讀取各類別的分店
        self.all_store_type_dict = { Type : {} for Type in self.store_type_list }
        for store_type, store_name in self.store_type_dict.items():
            for name in store_name:
                if name in self.final_data_dict.keys():
                    df = self.final_data_dict[name]
                    self.all_store_type_dict[store_type][name] = df
    
        
        self.all_type_sum_dict = {}
        for store_type, store_data_dict in self.all_store_type_dict.items():
            df = self._cal_total_sum_df(store_data_dict)
            self.all_type_sum_dict[store_type] = df
            
        
        return self.all_type_sum_dict
    
    #### 將所有店的合計與各個類別的店合計組成一個表
    def concat_all_type_store_sum(self):
        for index in self.total_df.index:
            #所有店
            all_store_df = pd.DataFrame(self.total_df.loc[index]).T
            all_store_df = convert_values_form(all_store_df)
            self.concat_all_type_store_df = pd.concat([self.concat_all_type_store_df, all_store_df])
            for types, df in self.all_type_sum_dict.items():
                type_store_df = pd.DataFrame(df.loc[index]).T
                type_store_df = convert_values_form(type_store_df)
                type_store_df = type_store_df.rename(index={index: index + f'({types})'})
                self.concat_all_type_store_df = pd.concat([self.concat_all_type_store_df, type_store_df])
                
        return self.concat_all_type_store_df  
    
    
    #### 計算同期單位淨利 和租金欄位一併插入 concat_all_type_store_df 中
    def _cal_same_period_profit(self):
        data_loader = DataLoader(self.Config, self.store_name_list, self.all_data_dict, center_kitchen_df = None)
        data_loader.get_file_data_three()
        self.same_period_profit_dict = data_loader.load_same_period_profit_dict()
        
        index_list = ['去年同期營業額', '今年同期營業額', '去年同期成長率']
        self.same_period_profit_df = pd.DataFrame(index=index_list, columns=self.concat_all_type_store_df.columns)
        for month, profit_dict in self.same_period_profit_dict.items():
            for key in profit_dict.keys():
                self.same_period_profit_df.loc[key, self.same_period_profit_df.columns[month-1]] = profit_dict[key]
        sum_df = self.same_period_profit_df.sum(axis=1)
        sum_df.loc['去年同期成長率'] = (sum_df.loc['去年同期營業額'] - sum_df.loc['今年同期營業額']) / sum_df.loc['去年同期營業額']
        self.same_period_profit_df.iloc[:, -1] = sum_df.values
        self.same_period_profit_df = convert_sam_period_values_form(self.same_period_profit_df) 
        
        position = self.concat_all_type_store_df.index.get_loc('單位淨利')
        # 把df分成上下兩部分
        upper_df = self.concat_all_type_store_df.iloc[:position+1]  # 包含單位淨利這行
        lower_df = self.concat_all_type_store_df.iloc[position+1:]  # 單位淨利之後
        
        # 把 new_df 插在中間
        self.concat_all_type_store_df = pd.concat([upper_df, self.same_period_profit_df, lower_df])
        
        #把租金欄位插在最下面
        self.concat_all_type_store_df = pd.concat([self.concat_all_type_store_df, self.rent_df])
        

                
        
# %% Class 中廚
class CenterKitchen(StoreSummary):   
    def __init__(self, Config):
        super().__init__(Config) 
        
        self.center_kitchen_df = None
        
    def build_CenterKitchen_all(self):
        #取得 -> center_kitchen_df
        self.get_CenterKitchen_excel_data_to_df()
        self.cal_other_centerkitchen_index_value()
        #取得 -> sum_include_ck_df
        self.sales_sum_include_center_kitchen()
        
        return self.center_kitchen_df, self.sum_include_ck_df
        
    def _build_ck_df(self):
        self.CenterKitchen_index_list = ["銷貨收入", "食材銷貨收入", "雜項銷貨收入", "公務費收入", "其他收入", "其他支出", "收入合計",
            "費用", "費用率", "費用率\n(不含公務費)", "實際毛利", "實際毛利率(%)", "單位淨利", "淨利率", "應發薪資",
            "薪資比", "奬金", "奬金比", "年終奬金",   "奬金比(年終)", "現場人員薪資", "工讀生薪資比\n(佔食材進貨金額)",
            "工時生產力", "薪資生產力"]
        self.columns_list = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十', '十一', '十二', '合  計']    
        # 建立空的 DataFrame
        df = pd.DataFrame(index=self.CenterKitchen_index_list, columns=self.columns_list)    
        
        return df
    
    #取得讀取excel檔的結果
    def get_CenterKitchen_excel_data_to_df(self):
        self.center_kitchen_df = self._build_ck_df()
        data_loader = DataLoader(self.Config, self.store_name_list, self.all_data_dict, center_kitchen_df = self.center_kitchen_df)

        self.center_kitchen_df = data_loader.load_center_kitchen_all()
        # return self.center_kitchen_df
    
    
    #計算其他index
    def cal_other_centerkitchen_index_value(self):
        #收入合計
        Calculator._cal_total_income(self.center_kitchen_df)
        #費用率 = 費用 / 收入合計    
        Calculator._cal_division(self.center_kitchen_df, '費用', '收入合計', '費用率')
        #費用率(不含公務費) = 費用 / 銷貨收入    
        Calculator._cal_division(self.center_kitchen_df, '費用', '銷貨收入', '費用率\n(不含公務費)')
        #淨利率 = 單位淨利 / 收入合計
        Calculator._cal_division(self.center_kitchen_df, '單位淨利', '收入合計', '淨利率')
        #薪資比 = 應發薪資 / 收入合計
        Calculator._cal_division(self.center_kitchen_df, '應發薪資', '收入合計', '薪資比')
        #獎金比 = 獎金 / 收入合計
        Calculator._cal_division(self.center_kitchen_df, '奬金', '收入合計', '獎金比')
        #奬金比(年終) = 年終獎金 / 收入合計
        Calculator._cal_division(self.center_kitchen_df, '年終奬金', '收入合計', '奬金比(年終)')
        
        self.center_kitchen_df = convert_center_kitchen_values_form(self.center_kitchen_df)
        
        return self.center_kitchen_df

 
    # 營業額總計(含中廚) + 單位淨利總計(含中廚)
    def sales_sum_include_center_kitchen(self):
        
        self.sum_include_ck_df = pd.DataFrame()
        self.sum_include_ck_df['營業額總計\n(含中廚)'] = self.center_kitchen_df.loc['收入合計'].add(self.total_df.loc['營業額'], fill_value=0)
        self.sum_include_ck_df['單位淨利總計\n(含中廚)'] = self.center_kitchen_df.loc['單位淨利'].add(self.total_df.loc['單位淨利'], fill_value=0)
        self.sum_include_ck_df = self.sum_include_ck_df.T
        self.sum_include_ck_df = convert_values_form(self.sum_include_ck_df)

        return self.sum_include_ck_df
        

    
    
# %% Class [ sheet_name = 114年度營業數據總表 ]

# =============================================================================
# SummaryReport: 負責計算營業數據總表（例如每月總計、平均、分段合計等）
# =============================================================================
class SummaryReport(StoreSummary):
    def __init__(self, Config):
        super().__init__(Config) 
        
        self.summary_df = None
    
    def build_SummaryReport_all(self):
        self._read_store_number()
        self.get_summary_df()
        self.cal_total_specicfy_month_sum_dict()
        
        
    #建立空的dataframe
    def _build_summary_df(self):
        # 定義 index
        self.index_summary_list = ['營業月', '營業額', '月平均營業額', '折扣金額', '折扣佔比', '銷貨收入', '費用', '費用率', 
                      '毛利', '毛利率', '單位淨利', '淨利率', '來客數', '客單價', '工時合計', '工時生產力']
        
        # 建立 column 列表 
        self.columns_list = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月', '總計']    
        # 建立空的 DataFrame
        df = pd.DataFrame(index=self.index_summary_list, columns=self.columns_list)    
        
        return df
    
    ### 讀取營業月 ###
    def _read_store_number(self):
        df = pd.read_excel(self.each_area_path, sheet_name = '各月店家數').iloc[:, 1:13]
        sum_value = list(df.sum(axis=1))
        df_to_list = df.values.tolist()[0]
        self.store_number_list = df_to_list + sum_value
        # return self.store_number_list
    
    #將各月合計資料取出輸入至數據總表   需計算之項目尚未進行計算
    def get_summary_df(self):
        # df = total_df
        self.summary_df = self._build_summary_df()
        
        #將店家數量輸入至summary_df
        self.summary_df.loc['營業月'] = self.store_number_list
                
        for summary_index in self.index_summary_list:
            if summary_index in self.total_df.index:
                row = self.total_df.loc[summary_index].values
                self.summary_df.loc[summary_index] = row
                
        #月平均營業額 = 營業額 / 營業月
        Calculator._cal_division(self.summary_df, '營業額', '營業月', '月平均營業額')  
        
        return self.summary_df
    
    ### 每三個月及六個月計算一次合計 ###
    @staticmethod
    def _cal_specicfy_month_sum(df, specicfy_month_number, specicfy_month_sum_dict):
        a = 1
        for i in range(0, 12, specicfy_month_number):
            specicfy_month_df = df.iloc[:, i:i+specicfy_month_number]
            specicfy_month_df_sum = pd.DataFrame(specicfy_month_df.sum(axis = 1))
            
            #月平均營業額 = 營業額 / 營業月
            Calculator._cal_division(specicfy_month_df_sum, '營業額', '營業月', '月平均營業額')             
            #折扣佔比 = 折扣金額 / 銷貨收入
            Calculator._cal_division(specicfy_month_df_sum, '折扣金額', '銷貨收入', '折扣佔比')  
            #費用率 = 費用 / 營業額
            Calculator._cal_division(specicfy_month_df_sum, '費用', '營業額', '費用率')  
            #毛利率 = 毛利 / 營業額
            Calculator._cal_division(specicfy_month_df_sum, '毛利', '營業額', '毛利率')  
            #淨利率 = 單位淨利 / 營業額
            Calculator._cal_division(specicfy_month_df_sum, '單位淨利', '營業額', '淨利率')  
            #客單價 = 營業額 / 來客數
            Calculator._cal_division(specicfy_month_df_sum, '營業額', '來客數', '客單價')  
            #工時生產力 = 營業額 / 工時合計
            Calculator._cal_division(specicfy_month_df_sum, '營業額', '工時合計', '工時生產力')  
            
            #設定columns
            if specicfy_month_number == 3:
                specicfy_month_df_sum.columns = ['小計']
            elif specicfy_month_number == 6:
                if a == 1:
                    specicfy_month_df_sum.columns = ['1~6月合計']
                elif a == 2:
                    specicfy_month_df_sum.columns = ['6~12月合計']
                    
            #輸入到dict中
            specicfy_month_sum_dict[specicfy_month_number][a] = specicfy_month_df_sum
            a = a + 1
            
        
    ### 完整的數據總表 ###    需計算之項目，計算完畢    
    def cal_total_specicfy_month_sum_dict(self):
        self.specicfy_month_sum_dict = {3:{}, 6:{}}
        
        #計算每三個月的合計
        SummaryReport._cal_specicfy_month_sum(self.summary_df, 3, self.specicfy_month_sum_dict)
        #計算每六個月的合計
        SummaryReport._cal_specicfy_month_sum(self.summary_df, 6, self.specicfy_month_sum_dict)
    
        #分別將其插入 summary_df 中
        for key, month_sum_dict in self.specicfy_month_sum_dict.items():
            insert_pos_list_3 = [3, 7, 11, 15]
            insert_pos_list_6 = [8, 17]
            for number, month_sum_df in month_sum_dict.items():
                if key == 3:
                    insert_pos = insert_pos_list_3[number-1]
                    self.summary_df = pd.concat([self.summary_df.iloc[:, :insert_pos], month_sum_df, self.summary_df.iloc[:, insert_pos:]], axis=1)
                elif key == 6:
                    insert_pos = insert_pos_list_6[number-1]
                    self.summary_df = pd.concat([self.summary_df.iloc[:, :insert_pos], month_sum_df, self.summary_df.iloc[:, insert_pos:]], axis=1)
        
        #將dataframe分為上下半部
        self.split_col = self.summary_df.columns.get_loc("1~6月合計")  # 找到 "1-6月合計" 的位置
        self.up_df = self.summary_df.iloc[:, :self.split_col + 1]  # 包含 "1-6月合計" 之前的部分
        self.down_df = self.summary_df.iloc[:, self.split_col + 1:]  # "7月" 之後的部分    
        self.up_df = convert_values_form(self.up_df)  #最終完整的dataframe 上半部
        self.down_df = convert_values_form(self.down_df)        
        
        return self.summary_df, self.up_df, self.down_df


#%% Class [ sheet_name = 總表 ]

class SimpleSummary(StoreSummary):
    def __init__(self, Config):
        super().__init__(Config)
        
    def _build_SimpleSummary_df(self):
        startwith_index = ['淨利率', '毛利率', '費用率', '薪資比', '工時生產力', '來客數', '客單價']
        startwith_index_list = []
        for index in startwith_index:
            result = [item for item in self.concat_all_type_store_df.index if item.startswith(index)]
            startwith_index_list.extend(result)   
            
        # 定義 index    
        index_SimpleSummary_list = ['總營業月', '營業目標', '營業目標達成率', '營業額', '平均月營業額', '成長率', '淨利目標', '單位淨利', '淨利達成率', '銷貨收入', '折扣金額',
                                    '折扣佔比'] + startwith_index_list + ['租金佔比(全)', '公積金', '營業額總計\n(含中廚)', '單位淨利總計\n(含中廚)', '淨利率(含中廚)']
        
        # 建立 column 列表 
        columns_list = ['營業店', '中廚']    
        # 建立空的 DataFrame
        df = pd.DataFrame(index=index_SimpleSummary_list, columns=columns_list)    

        return df
        
    def get_SimpleSummary_df(self):
        self.SimpleSummary_df = self._build_SimpleSummary_df()
        
        #[營業店]
        #從 concat_all_type_store_df 讀取資料
        self._get_df_data(self.concat_all_type_store_df)
        #從 sum_include_ck_df 讀取資料
        self._get_df_data(self.sum_include_ck_df)
        #淨利率(含中廚)
        self._cal_division('淨利率(含中廚)', '單位淨利總計\n(含中廚)', '營業額總計\n(含中廚)', '營業店')
        val = self.SimpleSummary_df.loc['淨利率(含中廚)', '營業店']
        if pd.notna(val):
            self.SimpleSummary_df.loc['淨利率(含中廚)', '營業店'] = f"%{val}"
        #總營業月
        self._get_business_month()
        #平均月營業額
        self._cal_division('平均月營業額', '營業額', '總營業月', '營業店')
        #公積金
        self._get_fund()
        
        #[中廚]
        #總營業月
        self.SimpleSummary_df.loc['總營業月', '中廚'] = 12
        #營業額(收入合計)
        self.SimpleSummary_df.loc['營業額', '中廚'] = self.center_kitchen_df.loc['收入合計'][-1]            
        #平均月營業額
        self._cal_division('平均月營業額', '營業額', '總營業月', '中廚')
        #成長率
        self.SimpleSummary_df.loc['成長率', '中廚'] = '自行計算' 
        #淨利率、毛利率、費用率、薪資比
        self._get_ck_data()


        
    #兩個值相除
    def _cal_division(self, result_index, up_index, down_index, col_index):
        self.SimpleSummary_df.loc[result_index, col_index] = self.SimpleSummary_df.loc[up_index, col_index] / self.SimpleSummary_df.loc[down_index, col_index]


    #[營業店]從df讀取資料
    def _get_df_data(self, df):
        for index in self.SimpleSummary_df.index:
            if index in df.index:
                self.SimpleSummary_df.loc[index, '營業店'] = df.loc[index][-1]
            
                
    #[營業店]輸入總營業月
    def _get_business_month(self):
        business_month_df = pd.read_excel(self.each_area_path, sheet_name = '各月店家數').iloc[:, 1:13]
        total_business_month = business_month_df.sum(axis=1)
        self.SimpleSummary_df.loc['總營業月', '營業店'] = total_business_month.values[0]
        
             
    #[營業店]公積金
    def _get_fund(self):
        df = pd.read_excel(self.this_year_path, sheet_name = '總表')
        row_idx, col_idx = df.eq('提列公積金(單位淨利7%)').to_numpy().nonzero()
        positions = list(zip(row_idx, col_idx))[0]   
        fund = df.iloc[positions[0], positions[1]+13]
        self.SimpleSummary_df.loc['公積金', '營業店'] = fund
    
    #[中廚]淨利率、毛利率、費用率、薪資比
    def _get_ck_data(self):
        index_list = ['淨利率', '毛利率', '費用率', '薪資比']
        for index in index_list:
            if index == '毛利率':
                self.SimpleSummary_df.loc[index, '中廚'] = self.center_kitchen_df.loc['實際毛利率(%)'][-1]
            else:
                self.SimpleSummary_df.loc[index, '中廚'] = self.center_kitchen_df.loc[index][-1]

        
        
    
#%% Function 轉換表現形式 -> 浮點數 or % 
#各家店    
def convert_values_form(df):
    convert_prcentage_index_list = ['折扣佔比','營業目標達成率','營業額佔比','成長率','費用率',
                                    '毛利率','實際毛利率','淨利達成率','淨利率','薪資比','奬金比']
    df = df.astype(float)
    for index in df.index:
        if index in convert_prcentage_index_list:
            df.loc[index] = df.loc[index].apply(lambda x: f"%{x}" if pd.notna(x) else x)
        elif index == '薪資生產力':
            df.loc[index] = df.loc[index].apply(lambda x: round(x, 2) if pd.notna(x) else x)
        elif index == '工時合計':
            df.loc[index] = df.loc[index].apply(lambda x: round(x, 1) if pd.notna(x) else x)
        else:
            df.loc[index] = df.loc[index].apply(lambda x: round(x, 0) if pd.notna(x) else x) 

    return df


#中廚
def convert_center_kitchen_values_form(df):
    convert_prcentage_index_list = ['費用率', "費用率\n(不含公務費)", "實際毛利率(%)", "淨利率", "薪資比"
                                    , "奬金比", "奬金比(年終)", "工讀生薪資比\n(佔食材進貨金額)"]
    df = df.astype(float)
    for index in df.index:
        if index in convert_prcentage_index_list:
            df.loc[index] = df.loc[index].apply(lambda x: f"%{x}" if pd.notna(x) else x)
        elif index == '薪資生產力':
            df.loc[index] = df.loc[index].apply(lambda x: round(x, 2) if pd.notna(x) else x)
        else:
            df.loc[index] = df.loc[index].apply(lambda x: round(x, 0) if pd.notna(x) else x) 

    return df

#同期單位淨利
def convert_sam_period_values_form(df):
    df = df.astype(float)
    for index in df.index:
        if index == '去年同期成長率':
            df.loc[index] = df.loc[index].apply(lambda x: f"%{x}" if pd.notna(x) else x)
        else:
            df.loc[index] = df.loc[index].apply(lambda x: round(x, 0) if pd.notna(x) else x) 

    return df


def convert_main_values_form(data_dict):
    
    convert_data_dict = OrderedDict()
    
    for name, df in data_dict.items():
        convert_df = convert_values_form(df)
        convert_data_dict[name] = convert_df   
        
    result_df = pd.concat(convert_data_dict)
    return result_df



# %% Class 將資料匯入excel 
def auto_thick_border(start_col=2):
    def decorator(func):
        def wrapper(self, ws, *args, **kwargs):
            # 先記錄原本的最大列
            start_row = ws.max_row + 1

            # 執行原本的寫入方法
            result = func(self, ws, *args, **kwargs)

            # 再記錄寫完後的最大列
            end_row = ws.max_row
            end_col = start_col + (ws.max_column - start_col)

            # 加上粗外框
            self._add_thick_border(ws, start_row, end_row, start_col, end_col)

            return result
        return wrapper
    return decorator
   
# =============================================================================
# ExcelReportWriter: 負責將結果寫入 Excel 檔
# =============================================================================
class ExcelReportWriter():
    def __init__(self, Config):
        self.wb = openpyxl.Workbook()    
        self.year = Config['year']
        self.thin_border = self._create_border('thin')
        self.thick_border = self._create_border('thick')
        
        
    def set_data_from(self, store_summary, center_kitchen, summary_report, simple_summary):
        self.store_name_list = store_summary.store_name_list
        self.index_list = store_summary.index_list
        self.total_df = store_summary.total_df
        self.final_df = convert_main_values_form(store_summary.final_data_dict)
        self.total_sales = store_summary.total_sales
        self.concat_all_type_store_df = store_summary.concat_all_type_store_df
        
        self.center_kitchen_df = center_kitchen.center_kitchen_df
        self.sum_include_ck_df = center_kitchen.sum_include_ck_df
        
        self.up_df = summary_report.up_df
        self.down_df = summary_report.down_df
        
        # if self.year == 12:
        self.SimpleSummary_df = simple_summary.SimpleSummary_df
        

    #細框及粗框
    @staticmethod
    def _create_border(style):
        return Border(
            left=Side(style=style, color='000000'),
            right=Side(style=style, color='000000'),
            top=Side(style=style, color='000000'),
            bottom=Side(style=style, color='000000'))

    
    #### sheet_name = 114年 ###
    def write_main_table(self):
        ws = self.wb.active
        ws.title = f"{self.year}年"
        
        # 寫入標題
        ws["A1"] = f"營業店{self.year}年度營業額及各項比率計算"
        ws.merge_cells("A1:O1")  # 合併標題儲存格
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(size=24)
        
        #寫入及合併第一欄 -> 店名列的cell
        self._write_store_name(ws)
        
        #寫入所有店家資料 -> final_df
        self._write_all_store_data(ws, self.final_df)
         
        #### 寫入最下面的各類型合計
        self._write_all_type_sum(ws)
   
        #### 寫入 營業額合計 + 單位淨利合計 欄位 
        self._write_sales_sum_col(ws)
               
        #### 寫入中廚 
        self._write_center_kitchen(ws, self.center_kitchen_df)
        
        #### 寫入營業額總計(含中廚) + 單位淨利總計(含中廚)
        self._write_all_type_sum_include_ck(ws, self.sum_include_ck_df)
    
        
    ####寫入及合併第一欄 -> 店名列的cell  
    def _write_store_name(self, ws):
        for i in range(len(self.store_name_list)):
            start_row = 3 + len(self.index_list) * i  
            end_row = 3 + len(self.index_list) * (i+1) -1 
            ws[f'A{start_row}'] = self.store_name_list[i]
            ws.merge_cells(f'A{start_row}:A{end_row}')
            ws[f'A{start_row}'].alignment = Alignment(horizontal='center', 
                                                      vertical='center', 
                                                      wrap_text=True)
            font = Font(name='微軟正黑體', size=30, bold=True)  # 假設你想要粗體，若不需要可移除 bold=True
            ws[f'A{start_row}'].font = font
        
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=1).border = self.thick_border  # column=1 表示 A 欄
        ws.column_dimensions['A'].width = 15
    
    #### 寫入所有店家資料 -> final_df
    def _write_all_store_data(self, ws, final_df):
        #寫入col
        for col_idx, column_name in enumerate(final_df.columns, 3):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = column_name
            cell.font = Font(name="微軟正黑體", size=22, color = '8B4513')
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        #畫col的外框
        for i in range(1, 16):
            cell = ws.cell(row=2, column=i)
            cell.border = self.thin_border
         
        # 寫入數值    
        final_df = final_df.reset_index(level = 1) #把index變成第一欄
        for row_idx, row in enumerate(final_df.values, 3):
            for col_idx, value in enumerate(row, 2): 
                # 設定字體顏色大小
                cell = self._set_color_font_border(ws, cell, col_idx, row_idx, value, '0000FF')                
                #設定黑色外框
                self._set_thick_border(cell, row, row_idx, col_idx)
                #設定數字格式
                ExcelReportWriter._set_value_format(value, cell)
                #設定百分比
                cell = ExcelReportWriter._apply_percentage_format(ws, cell, value)
         
        #調整欄寬
        ExcelReportWriter._adjust_col_width(ws, 40)
        # 調整行高
        ExcelReportWriter._adjust_row_height(ws, 30)
    
    
    
    #### 寫入最下面的各類型合計
    @auto_thick_border()
    def _write_all_type_sum(self, ws):
        max_row = ws.max_row + 1
        concat_all_type_store_reset_index = self.concat_all_type_store_df.reset_index() #把項目變成第一欄
  
        for row_idx, row in enumerate(concat_all_type_store_reset_index.values, max_row):
            for col_idx, value in enumerate(row, 2):  # 從第2欄開始寫  
                cell = ws.cell(row=row_idx, column=col_idx)       
                cell = self._set_color_font_border(ws, cell, col_idx, row_idx, value, "FF0000")
                #設定數字格式
                ExcelReportWriter._set_value_format(value, cell)
                #設定百分比
                cell = ExcelReportWriter._apply_percentage_format(ws, cell, value)
        #畫合計欄位(第一欄)
        self._write_sum_col(ws, concat_all_type_store_reset_index, max_row, '合計')
    
    #畫合計欄位(第一欄)
    def _write_sum_col(self, ws, df, max_row, name):
        start_row = max_row
        end_row = max_row + len(df.index)-1
        ws[f'A{start_row}'] = name
        ws.merge_cells(f'A{start_row}:A{end_row}')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center', 
                                                  vertical='center', 
                                                  wrap_text=True)
        font = Font(name='微軟正黑體', size=30, bold=True)  
        ws[f'A{start_row}'].font = font
        for row in range(start_row, end_row+1):
            ws.cell(row=row, column=1).border = self.thick_border  # column=1 表示 A 欄

    
    #### 寫入 營業額合計 + 單位淨利合計 欄位
    @auto_thick_border(start_col=1)
    def _write_sales_sum_col(self, ws):
        sales_sum = int(self.concat_all_type_store_df.loc['營業額'][-1])
        sales_sum_format = f"{sales_sum:,}"
        self._set_all_year_total_col(ws, '營業額', sales_sum_format, "FFFFCC")
                  
        #單位淨利合計
        unit_net_profit = int(self.concat_all_type_store_df.loc['單位淨利'][-1])  
        unit_net_profit_format = f"{unit_net_profit:,}"
        self._set_all_year_total_col(ws, '單位淨利', unit_net_profit_format, "FFFFCC")
        
        #淨利率
        net_profit_ratio = pd.Series(unit_net_profit / sales_sum).apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else x)[0]
        self._set_all_year_total_col(ws, '淨利率', net_profit_ratio, "FFFFCC")
     
        
    #營業額合計 + 單位淨利合計 欄位設定
    def _set_all_year_total_col(self, ws, index, value, color):
        merge_row = ws.max_row + 1
        # 設定該列高度為 40
        ws.row_dimensions[merge_row].height = 50        
        # 合併儲存格
        ws.merge_cells(f"A{merge_row}:O{merge_row}")        
        # 設定儲存格內容與樣式
        cell = ws.cell(row=merge_row, column=1)   
        cell.value = f'{self.year}年度{index}合計 : {value}'       
        # 設定填滿色與樣式
        light_yellow_fill = PatternFill(fill_type="solid", fgColor=color)
        cell.fill = light_yellow_fill
        cell.font = Font(name="微軟正黑體", size=22, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")    
    
        
    #### 寫入中廚 
    @auto_thick_border()
    def _write_center_kitchen(self, ws, center_kitchen_df):
        max_row = ws.max_row + 1
        center_kitchen_reset_index_df = center_kitchen_df.reset_index() #把項目變成第一欄
    
        
        for row_idx, row in enumerate(center_kitchen_reset_index_df.values, max_row):
            for col_idx, value in enumerate(row, 2):  # 從第2欄開始寫 
                cell = ws.cell(row=row_idx, column=col_idx)             
                cell = self._set_color_font_border(ws, cell, col_idx, row_idx, value, "0000FF")
                #設定數字格式
                ExcelReportWriter._set_value_format(value, cell)
                #設定百分比
                cell = ExcelReportWriter._apply_percentage_format(ws, cell, value)
        #畫中廚欄位(第一欄)
        self._write_sum_col(ws, center_kitchen_reset_index_df, max_row, '中廚')


    #### 寫入營業額總計(含中廚) + 單位淨利總計(含中廚)
    def _write_all_type_sum_include_ck(self, ws, sum_include_ck_df):
        max_row = ws.max_row + 1
        
        for row_idx, value in enumerate(sum_include_ck_df.index, max_row):
            cell = ws.cell(row=row_idx, column=1)   
            cell.value = value
            cell.font = Font(name="微軟正黑體", size=16, color='0000FF')
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
            light_yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
            cell.fill = light_yellow_fill
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        
        for row_idx, row in enumerate(sum_include_ck_df.values, max_row):
            for col_idx, value in enumerate(row, 3):  # 從第3欄開始寫  
                cell = ws.cell(row=row_idx, column=col_idx)
                cell = self._set_color_font_border(ws, cell, col_idx, row_idx, value, "0000FF")
                light_yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
                cell.fill = light_yellow_fill
                # 設定數字格式
                ExcelReportWriter._set_value_format(value, cell)
        self._sales_sum_include_centerkitchen_col(ws, sum_include_ck_df)          
    
        
    #### 營業額合計(含中廚) + 單位淨利合計(含中廚) 欄位設定
    @auto_thick_border(start_col=1)
    def _sales_sum_include_centerkitchen_col(self, ws, df):
        sales_sum = int(df.loc['營業額總計\n(含中廚)'][-1])
        sales_sum_format = f"{sales_sum:,}"
        self._set_all_year_total_col(ws, '營業額總計(含中廚)', sales_sum_format, 'CCFFCC')
                  
        #單位淨利合計
        unit_net_profit = int(df.loc['單位淨利總計\n(含中廚)'][-1])  
        unit_net_profit_format = f"{unit_net_profit:,}"
        self._set_all_year_total_col(ws, '單位淨利總計(含中廚)', unit_net_profit_format, 'CCFFCC')
        
        #淨利率
        net_profit_ratio = pd.Series(unit_net_profit / sales_sum).apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else x)[0]
        self._set_all_year_total_col(ws, '淨利率(含中廚)', net_profit_ratio, 'CCFFCC')
    
    
    
    #### 設定字體、顏色、大小、外框
    def _set_color_font_border(self, ws, cell, col_index, row_index, value, color):
        cell = ws.cell(row=row_index, column=col_index)
        cell.value = value
        cell.border = self.thin_border
        if col_index == 2:  #項目
           cell.font = Font(name="微軟正黑體", size=16, color=color)
           cell.alignment = Alignment(horizontal="left", vertical="center")
        else:  #數值
            cell.font = Font(name="Arial Narrow", size=22)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return cell
    
    #### 設定 final_df 黑色外框
    def _set_thick_border(self, cell, row, row_idx, col_idx):
        # 確定最右邊的column index (df有幾欄就加到幾)
        max_col_idx = len(row) + 1  # 因為col_idx從2開始,所以+1    
        # 判斷是否需要畫粗外框
        if (row_idx - 3) % len(self.index_list) == 0 and col_idx == 2:  # 每組第一行的第一個cell
            cell.border = Border(top=self.thick_border.top, left=self.thick_border.left, bottom=self.thin_border.bottom, right=self.thin_border.right)
        elif (row_idx - 3) % len(self.index_list) == 0 and col_idx == max_col_idx:  # 每組第一行的最右cell
            cell.border = Border(top=self.thick_border.top, right=self.thick_border.right, bottom=self.thin_border.bottom, left=self.thin_border.left)
        elif (row_idx - 3) % len(self.index_list) == 0:  # 每組第一行的其他cell
            cell.border = Border(top=self.thick_border.top, bottom=self.thin_border.bottom, left=self.thin_border.left, right=self.thin_border.right)
        elif (row_idx - 2) % len(self.index_list) == 0 and col_idx == 2:  # 每組最後一行的第一個cell
            cell.border = Border(bottom=self.thick_border.bottom, left=self.thick_border.left, top=self.thin_border.top, right=self.thin_border.right)
        elif (row_idx - 2) % len(self.index_list) == 0 and col_idx == max_col_idx:  # 每組最後一行的最右cell
            cell.border = Border(bottom=self.thick_border.bottom, right=self.thick_border.right, top=self.thin_border.top, left=self.thin_border.left)
        elif (row_idx - 2) % len(self.index_list) == 0:  # 每組最後一行的其他cell
            cell.border = Border(bottom=self.thick_border.bottom, top=self.thin_border.top, left=self.thin_border.left, right=self.thin_border.right)
        elif col_idx == 2:  # 中間行的第一個cell
            cell.border = Border(left=self.thick_border.left, top=self.thin_border.top, bottom=self.thin_border.bottom, right=self.thin_border.right)
        elif col_idx == max_col_idx:  # 中間行的最右cell
            cell.border = Border(right=self.thick_border.right, top=self.thin_border.top, bottom=self.thin_border.bottom, left=self.thin_border.left)
        else:  # 其他cell維持原有的thin_border
            cell.border = self.thin_border
            
            
            
    
    #### sheet_name = 114年度營業數據總表
    def write_sales_summary_table(self, wb):
        ws = self.wb.create_sheet(f"{self.year}年度營業數據總表")
        
        # 寫入標題
        ws["A1"] = f"{self.year}年度營業額數據"
        ws.merge_cells("A1:K1")  # 合併標題儲存格
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(bold=True, size=24)
        
        ### 寫入上半部 ###
        self._write_up_df(ws)
                                
        ### 寫入下半部 ###
        self._write_down_df(ws)
        
        #調整欄寬
        ExcelReportWriter._adjust_col_width(ws, 25)
        # 調整行高
        ExcelReportWriter._adjust_row_height(ws, 30)
            
        #繪製粗框
        for i in range(2):
            start_row = 2 + (len(self.up_df.index))*i + i
            end_row = 2 + (len(self.up_df.index))*(i+1) + i
            start_col = 1 
            end_col = 11   
            self._add_thick_border(ws, start_row, end_row, start_col, end_col)
    
    def _write_up_df(self, ws):
        #寫入col
        up_col = ['月份'] + list(self.up_df.columns)
        for col_idx, column_name in enumerate(up_col, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = column_name
            cell.font = Font(name="微軟正黑體", size=18, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        #寫入index
        for row_idx, row_name in enumerate(self.up_df.index, 3):
            cell = ws.cell(row=row_idx, column=1)
            cell.value = row_name
            cell.font = Font(name="微軟正黑體", size=18, color="0000FF")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            
        #寫入數值
        for row_idx, row in enumerate(self.up_df.values, 3):
            for col_idx, value in enumerate(row, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                #設定百分比
                cell = ExcelReportWriter._apply_percentage_format(ws, cell, value)
                cell.font = Font(name="微軟正黑體", size=16)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # 如果是數字則設定格式
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0"  # 千位逗號分隔，無小數點
    
    
    def _write_down_df(self, ws):
        #寫入col
        down_col = ['月份'] + list(self.down_df.columns)
        for col_idx, column_name in enumerate(down_col, 1):
            cell = ws.cell(row=2+len(self.up_df.index)+1, column=col_idx)
            cell.value = column_name
            cell.font = Font(name="微軟正黑體", size=18, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        #寫入index
        for row_idx, row_name in enumerate(self.down_df.index, 3+len(self.up_df.index)+1):
            cell = ws.cell(row=row_idx, column=1)
            cell.value = row_name
            cell.font = Font(name="微軟正黑體", size=18, color="0000FF")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        #寫入數值
        for row_idx, row in enumerate(self.down_df.values, 3+len(self.up_df.index)+1):
            for col_idx, value in enumerate(row, 2):  
                cell = ws.cell(row=row_idx, column=col_idx)
                #設定百分比
                cell = ExcelReportWriter._apply_percentage_format(ws, cell, value)
                cell.font = Font(name="微軟正黑體", size=16)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # 如果是數字則設定格式
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0"  # 千位逗號分隔，無小數點
    

    def _add_thick_border(self, ws, start_row, end_row, start_col, end_col):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = self.thin_border
    
        for col in range(start_col, end_col + 1):
            # 上邊
            cell = ws.cell(row=start_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=Side(style='thick', color='000000'),
                bottom=cell.border.bottom
            )
            # 下邊
            cell = ws.cell(row=end_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=Side(style='thick', color='000000')
            )
    
        for row in range(start_row, end_row + 1):
            # 左邊
            cell = ws.cell(row=row, column=start_col)
            cell.border = Border(
                left=Side(style='thick', color='000000'),
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            # 右邊
            cell = ws.cell(row=row, column=end_col)
            cell.border = Border(
                left=cell.border.left,
                right=Side(style='thick', color='000000'),
                top=cell.border.top,
                bottom=cell.border.bottom
            )
    

    #### sheet_name = 總表
    def write_simple_summary_table(self, wb):
        ws = self.wb.create_sheet("總表")
        
        # 合併儲存格
        ws.merge_cells("B2:C2")
        # 寫入標題
        ws["B2"] = f"{self.year}"
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].font = Font(name="微軟正黑體", size=20, color='0000FF')
        ws["B2"].border = self.thin_border
        # 對 B2:C2 中的每個儲存格都加上邊框
        for col in range(2, 4):  # B 是第2欄，C是第3欄
            cell = ws.cell(row=2, column=col)
            cell.border = self.thin_border

        
        for row_idx, value in enumerate(self.SimpleSummary_df.index, 4):
            cell = ws.cell(row=row_idx, column=1)   
            cell.value = value
            cell.font = Font(name="微軟正黑體", bold=True, size=16, color='0000FF')
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
            
        for col_idx, value in enumerate(self.SimpleSummary_df.columns, 2):
            cell = ws.cell(row=3, column=col_idx)   
            cell.value = value
            cell.font = Font(name="微軟正黑體", size=20, color='0000FF')
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border
            
        for row_idx, row in enumerate(self.SimpleSummary_df.values, 4):
            for col_idx, value in enumerate(row, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                #設定數字格式
                ExcelReportWriter._set_value_format(value, cell)
                #設定百分比
                cell = ExcelReportWriter._apply_percentage_format(ws, cell, value)
                
                cell.font = Font(name="Arial Narrow", size=20)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.thin_border
                
                
        #合併儲存格
        for i in range(1, 5):
            #淨利率, 毛利率, 費用率, 薪資比
            ws.merge_cells(start_row=16+(i-1)*3, start_column=3, end_row=16+i*3-1, end_column=3)
        for i in range(5):
            ws.merge_cells(start_row=37+i, start_column=2, end_row=37+i, end_column=3)
            
        
        
        #調整欄寬
        ExcelReportWriter._adjust_col_width(ws, 40)
        # 調整行高
        ExcelReportWriter._adjust_row_height(ws, 30)
    
    #%%小工具    
    # 調整欄寬
    @staticmethod
    def _adjust_col_width(ws, col_width):
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = col_width
    # 調整行高
    @staticmethod
    def _adjust_row_height(ws, row_height):        
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = row_height    

    # 如果是數字則設定格式
    @staticmethod
    def _set_value_format(value, cell):
        if isinstance(value, (int, float)):
            if isinstance(value, float) and not value.is_integer():
                cell.number_format = "#,##0.##"  # 顯示小數點（最多 2 位，不足不補零）
            else:
                cell.number_format = "#,##0"  # 整數格式，無小數點               
    
    #設定為百分比
    @staticmethod
    def _apply_percentage_format(ws, cell, val):
        if isinstance(val, str) and val.startswith('%'):
            float_val = float(val[1:])
            cell.value = float_val
            cell.number_format = '0.00%'  # 設定百分比格式
        else:
            cell.value = val
        return cell
        
        


# %% class 主程式執行
# =============================================================================
# ReportController: 協調各部件，執行整個報表流程
# =============================================================================
class ReportCoordinator:
    def __init__(self, config):
        self.config = config
        self.year = config['year']
        self.output_folder_path = config['output_folder_path']
        self.store_summary = StoreSummary(config)
        self.center_kitchen = CenterKitchen(config)
        self.summary_report = SummaryReport(config)
        self.simple_summary = SimpleSummary(config)
        self.writer = ExcelReportWriter(config)

    def prepare_data(self):
        print("🔵 正在建立所有店家年度報表...")
        self.store_summary.build_StoreSummary_all()

        print("🔵 正在讀取中央廚房資料...")
        self.center_kitchen.total_df = self.store_summary.total_df
        self.center_kitchen.build_CenterKitchen_all()

        print("🔵 正在建立營業總表...")
        self.summary_report.total_df = self.store_summary.total_df
        self.summary_report.build_SummaryReport_all()
        
        print("🔵 正在建立總表...")  
        self.simple_summary.center_kitchen_df = self.center_kitchen.center_kitchen_df
        self.simple_summary.sum_include_ck_df = self.center_kitchen.sum_include_ck_df
        self.simple_summary.concat_all_type_store_df = self.store_summary.concat_all_type_store_df       
        self.simple_summary.get_SimpleSummary_df()
        

    def write_excel(self):
        print("🟢 正在寫入Excel...")
        # 把資料注入 writer
        self.writer.set_data_from(self.store_summary, self.center_kitchen, self.summary_report, self.simple_summary)

        self.writer.write_main_table()
        self.writer.write_sales_summary_table(self.writer.wb)
        self.writer.write_simple_summary_table(self.writer.wb)
        output_path = self.output_folder_path + f'/營業店年度營業額及各項比率計算({self.year}年).xlsx'
        self.writer.wb.save(output_path)
        print(f"✅ 完成！報表儲存到：{output_path}")

    def run_all(self):
        self.prepare_data()
        self.write_excel()

